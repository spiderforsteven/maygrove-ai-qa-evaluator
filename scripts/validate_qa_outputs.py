#!/usr/bin/env python3
"""
Validation script for maygrove-ai-qa-evaluator outputs.
Run after generating Q&A sets to catch structural issues before delivery.

Usage:
    python scripts/validate_qa_outputs.py /path/to/output/dir

Checks:
    - JSONL files are valid JSON
    - Every verified item has required fields
    - All intents and personas are from the allowed sets
    - Every fact has 'text' and 'source' keys
    - Confidence values are in [0, 1]
    - IDs are unique
    - CSV is readable and has expected columns
    - Excel is readable and has expected sheets (if present)
"""

import json
import sys
import csv
from pathlib import Path

EXPECTED_INTENTS = {
    "PlantingTech", "ProductInquiry", "GrowingPlan", "Lifestyle",
    "ChatEmotional", "DietCooking", "OutofScope"
}

EXPECTED_PERSONAS = {
    "Newbie", "Busy Professional", "Green Enthusiast", "Troubleshooter"
}

REQUIRED_FIELDS = {
    "id", "intent", "persona", "scenario", "question", "reference_answer",
    "facts", "confidence", "tags", "verification_status", "verification_confidence",
    "verification_notes", "issues", "source_gap"
}

CSV_COLUMNS = [
    "id", "intent", "persona", "scenario", "question", "reference_answer",
    "facts_count", "source_sections", "confidence", "verification_status",
    "verification_confidence", "source_gap", "tags"
]


def validate_jsonl(filepath: Path) -> list[dict]:
    """Parse JSONL and return list of records; raises ValueError on invalid JSON."""
    records = []
    with open(filepath, "r", encoding="utf-8") as f:
        for line_no, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as exc:
                raise ValueError(f"{filepath}:{line_no}: invalid JSON: {exc}") from exc
    return records


def validate_record(record: dict, line_no: int) -> list[str]:
    errors = []
    missing = REQUIRED_FIELDS - set(record.keys())
    if missing:
        errors.append(f"line {line_no}: missing fields {sorted(missing)}")

    intent = record.get("intent")
    if intent and intent not in EXPECTED_INTENTS:
        errors.append(f"line {line_no}: unexpected intent '{intent}'")

    persona = record.get("persona")
    if persona and persona not in EXPECTED_PERSONAS:
        errors.append(f"line {line_no}: unexpected persona '{persona}'")

    confidence = record.get("confidence")
    if confidence is not None and not (0 <= float(confidence) <= 1):
        errors.append(f"line {line_no}: confidence {confidence} out of range")

    verification_confidence = record.get("verification_confidence")
    if verification_confidence is not None and not (0 <= float(verification_confidence) <= 1):
        errors.append(f"line {line_no}: verification_confidence {verification_confidence} out of range")

    facts = record.get("facts")
    if facts is not None:
        if not isinstance(facts, list):
            errors.append(f"line {line_no}: 'facts' should be a list")
        else:
            for idx, fact in enumerate(facts):
                if not isinstance(fact, dict):
                    errors.append(f"line {line_no}: fact[{idx}] is not an object")
                    continue
                if "text" not in fact or "source" not in fact:
                    errors.append(f"line {line_no}: fact[{idx}] missing 'text' or 'source'")

    return errors


def validate_csv(filepath: Path) -> list[str]:
    errors = []
    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames != CSV_COLUMNS:
            errors.append(f"{filepath}: unexpected columns {reader.fieldnames}")
        row_count = sum(1 for _ in reader)
    print(f"  {filepath}: {row_count} rows")
    return errors


def validate_excel(filepath: Path) -> list[str]:
    errors = []
    try:
        import openpyxl
    except ImportError as exc:
        errors.append(f"{filepath}: openpyxl not installed, cannot validate Excel")
        return errors

    wb = openpyxl.load_workbook(filepath, read_only=True)
    expected_sheets = {"Verified QA", "Verification Report", "Rejected"}
    missing = expected_sheets - set(wb.sheetnames)
    if missing:
        errors.append(f"{filepath}: missing sheets {sorted(missing)}")
    else:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row_count = sum(1 for _ in ws.iter_rows())
            print(f"  {filepath} / {sheet}: {row_count} rows")
    return errors


def main() -> int:
    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} <output_dir>")
        return 1

    output_dir = Path(sys.argv[1])
    if not output_dir.is_dir():
        print(f"Error: {output_dir} is not a directory")
        return 1

    all_errors = []

    jsonl_files = [
        ("verified_qa.jsonl", True),
        ("cross_verification_report.jsonl", False),
        ("rejected_needs_review.jsonl", False),
    ]
    for name, strict in jsonl_files:
        filepath = output_dir / name
        if not filepath.exists():
            print(f"  {filepath}: not found (skipping)")
            continue
        try:
            records = validate_jsonl(filepath)
        except ValueError as exc:
            all_errors.append(str(exc))
            continue
        print(f"  {filepath}: {len(records)} records")

        seen_ids = set()
        for line_no, record in enumerate(records, start=1):
            record_id = record.get("id")
            if record_id:
                if record_id in seen_ids:
                    all_errors.append(f"{filepath}: duplicate id '{record_id}'")
                seen_ids.add(record_id)
            if strict:
                all_errors.extend(validate_record(record, line_no))

    csv_path = output_dir / "verified_qa.csv"
    if csv_path.exists():
        all_errors.extend(validate_csv(csv_path))
    else:
        print(f"  {csv_path}: not found (skipping)")

    excel_path = output_dir / "verified_qa.xlsx"
    if excel_path.exists():
        all_errors.extend(validate_excel(excel_path))
    else:
        print(f"  {excel_path}: not found (skipping)")

    if all_errors:
        print(f"\nFound {len(all_errors)} error(s):")
        for err in all_errors:
            print(f"  - {err}")
        return 1

    print("\nAll validation checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
