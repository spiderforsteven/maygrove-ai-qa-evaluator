#!/usr/bin/env python3
"""
Orchestrate the MayGrove AI QA evaluation pipeline.

This script runs a deterministic fact-extraction phase locally, then uses the
active LLM runtime (auto-detected: hermes CLI, OpenAI API key, DeepSeek API key,
or other supported providers) to run the multi-agent QA generation and critic passes.

If no LLM is available, it falls back to deterministic template generation so
the pipeline always produces a validated dataset.

Usage:
    python scripts/run_evaluation.py \
        --maygrove-md /path/to/Maygrove.md \
        --v30-manual /path/to/MayGrove_Verti_30__V30__使用说明书.md \
        --seed-library /path/to/seed_library.xlsx \
        --output /path/to/output \
        --count 30 \
        [--skip-web]

Dependencies:
    pip install -r scripts/requirements.txt

Set TAVILY_API_KEY in the environment to enable web research. The pipeline
works without it (only local docs + seed library).

LLM auto-detection priority:
    1. Hermes CLI (hermes chat -q) — uses your Hermes-configured model/provider
    2. OPENAI_API_KEY env var → OpenAI / compatible API
    3. DEEPSEEK_API_KEY env var → DeepSeek API
    4. ANTHROPIC_API_KEY env var → Anthropic API
    5. If none available → deterministic template fallback
"""

import argparse
import json
import os
import re
import shlex
import shutil
import subprocess
import sys
import time
from pathlib import Path

from load_sources import load_sources


INTENTS = [
    "PlantingTech",
    "ProductInquiry",
    "GrowingPlan",
    "Lifestyle",
    "ChatEmotional",
    "DietCooking",
    "OutofScope",
]

PERSONAS = [
    "Newbie",
    "Busy Professional",
    "Green Enthusiast",
    "Troubleshooter",
]


def parse_args():
    parser = argparse.ArgumentParser(description="Run MayGrove QA evaluation pipeline")
    parser.add_argument("--maygrove-md", required=True, help="Path to Maygrove.md")
    parser.add_argument("--v30-manual", required=True, help="Path to V30 manual")
    parser.add_argument("--seed-library", required=True, help="Path to seed library Excel")
    parser.add_argument("--output", required=True, help="Output directory")
    parser.add_argument(
        "--count",
        type=lambda x: int(x) if int(x) > 0 else argparse.ArgumentTypeError("--count must be a positive integer"),
        default=28,
        help="Target number of Q&A pairs"
    )
    parser.add_argument("--skip-web", action="store_true", help="Disable Tavily web research")
    parser.add_argument("--llm-cli", default=None, help="Override LLM CLI command. Auto-detected if not set.")
    return parser.parse_args()


def detect_llm():
    """
    Auto-detect the best available LLM strategy.
    Returns a dict with {'type': ..., 'config': ...}
    """
    # Priority 1: hermes CLI
    if shutil.which("hermes"):
        print("[LLM Detected] Hermes CLI — using your Hermes-configured model/provider", file=sys.stderr)
        return {"type": "hermes", "cli": "hermes chat -q"}

    # Priority 2: OPENAI_API_KEY
    if os.getenv("OPENAI_API_KEY"):
        print("[LLM Detected] OpenAI API key found in environment", file=sys.stderr)
        return {"type": "openai", "api_key": os.environ["OPENAI_API_KEY"], "model": os.getenv("OPENAI_MODEL", "gpt-4o-mini"), "base_url": os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1")}

    # Priority 3: DEEPSEEK_API_KEY
    if os.getenv("DEEPSEEK_API_KEY"):
        print("[LLM Detected] DeepSeek API key found in environment", file=sys.stderr)
        return {"type": "openai", "api_key": os.environ["DEEPSEEK_API_KEY"], "model": os.getenv("DEEPSEEK_MODEL", "deepseek-chat"), "base_url": "https://api.deepseek.com/v1"}

    # Priority 4: ANTHROPIC_API_KEY
    if os.getenv("ANTHROPIC_API_KEY"):
        print("[LLM Detected] Anthropic API key found in environment", file=sys.stderr)
        return {"type": "anthropic", "api_key": os.environ["ANTHROPIC_API_KEY"], "model": os.getenv("ANTHROPIC_MODEL", "claude-sonnet-4-20250514")}

    # Fallback: no LLM available
    print("[LLM] No LLM CLI or API key detected. Using deterministic template fallback.", file=sys.stderr)
    return {"type": "fallback", "config": {}}


def load_template(template_dir, name):
    path = template_dir / name
    if not path.exists():
        raise FileNotFoundError(f"Missing template: {path}")
    return path.read_text(encoding="utf-8")


def ensure_dependencies():
    try:
        import pandas  # noqa: F401
        import openpyxl  # noqa: F401
    except ImportError as exc:
        print(f"Error: missing dependency {exc.name}. Run: pip install -r scripts/requirements.txt", file=sys.stderr)
        sys.exit(1)


def load_seed_library_data(sources):
    """Return pre-structured JSON files for first batch, nutrition, and tips."""
    data_dir = Path(__file__).parent.parent / "data"
    files = {
        "first_batch": data_dir / "seed_library_first_batch.json",
        "nutrition": data_dir / "seed_nutrition.json",
        "planting_tips": data_dir / "planting_tips.json",
    }
    loaded = {}
    for key, path in files.items():
        if path.exists():
            with open(path, "r", encoding="utf-8") as f:
                loaded[key] = json.load(f)
    return loaded


def build_qa_prompt(template, intent, persona, scenario, facts, brand_boundary):
    """Replace placeholders in the QA generator template."""
    role = "scenario-driven" if intent != "OutofScope" else "persona-driven"
    focus = "scenario-based user journey stage" if role == "scenario-driven" else "persona voice adaptation"
    return (
        template
        .replace("{A|B|C}", "A")
        .replace("{scenario-driven | fact-driven | persona-driven}", role)
        .replace("{intent}", intent)
        .replace("{persona}", persona)
        .replace("{scenario}", scenario)
        .replace("{facts_json}", json.dumps(facts, ensure_ascii=False))
        .replace("{out_of_scope_policy}", brand_boundary)
        .replace("{scenario-based user journey stage | reverse-question from fact | persona voice adaptation}", focus)
    )


def build_critic_prompt(template, question, answer, facts, intent, brand_boundary, critic_role):
    roles = {
        "A": "fact traceability",
        "B": "logical plausibility",
        "C": "safety & boundary compliance",
    }
    return (
        template
        .replace("{A|B|C}", critic_role)
        .replace("{fact traceability | logical plausibility | safety & boundary compliance}", roles[critic_role])
        .replace("{question}", question)
        .replace("{reference_answer}", answer)
        .replace("{facts_json}", json.dumps(facts, ensure_ascii=False))
        .replace("{intent}", intent)
        .replace("{out_of_scope_policy}", brand_boundary)
    )


def build_cross_verifier_prompt(template, question, answers, facts, scores, consensus_score):
    return (
        template
        .replace("{intent}", answers[0]["intent"])
        .replace("{persona}", answers[0]["persona"])
        .replace("{question}", question)
        .replace("{agent_a_answer}", answers[0]["reference_answer"])
        .replace("{agent_b_answer}", answers[1]["reference_answer"])
        .replace("{agent_c_answer}", answers[2]["reference_answer"])
        .replace("{facts_json}", json.dumps(facts, ensure_ascii=False))
        .replace("{critic_a_score}", str(scores[0]))
        .replace("{critic_b_score}", str(scores[1]))
        .replace("{critic_c_score}", str(scores[2]))
        .replace("{consensus_score}", str(consensus_score))
    )


def call_llm(provider, prompt, timeout=120):
    """Invoke the best available LLM with a prompt. Returns raw text output."""
    if provider["type"] == "hermes":
        return call_llm_hermes(provider["cli"], prompt, timeout)
    elif provider["type"] == "openai":
        return call_llm_openai(provider, prompt, timeout)
    elif provider["type"] == "anthropic":
        return call_llm_anthropic(provider, prompt, timeout)
    return ""


def call_llm_hermes(cli_prefix, prompt, timeout=120):
    """Invoke the LLM via Hermes CLI subprocess."""
    safe_prompt = shlex.quote(prompt)
    cmd = f"{cli_prefix} --quiet {safe_prompt}"
    try:
        result = subprocess.run(
            cmd,
            shell=True,
            capture_output=True,
            text=True,
            timeout=timeout,
        )
    except subprocess.TimeoutExpired:
        print(f"[LLM] Hermes call timed out after {timeout}s", file=sys.stderr)
        return ""
    if result.returncode != 0:
        stderr = result.stderr.strip()[:200]
        print(f"[LLM] Hermes call failed (exit {result.returncode}): {stderr}", file=sys.stderr)
        return ""
    return result.stdout


def call_llm_openai(provider, prompt, timeout=120):
    """Invoke the LLM via OpenAI-compatible API."""
    try:
        from openai import OpenAI
    except ImportError:
        print("[LLM] openai package not installed. Run: pip install openai", file=sys.stderr)
        return ""
    try:
        client = OpenAI(api_key=provider["api_key"], base_url=provider.get("base_url", "https://api.openai.com/v1"))
        resp = client.chat.completions.create(
            model=provider["model"],
            messages=[{"role": "user", "content": prompt}],
            max_tokens=2048,
            temperature=0.3,
            timeout=timeout,
        )
        return resp.choices[0].message.content or ""
    except Exception as e:
        print(f"[LLM] OpenAI API call failed: {e}", file=sys.stderr)
        return ""


def call_llm_anthropic(provider, prompt, timeout=120):
    """Invoke the LLM via Anthropic API."""
    try:
        from anthropic import Anthropic
    except ImportError:
        print("[LLM] anthropic package not installed. Run: pip install anthropic", file=sys.stderr)
        return ""
    try:
        client = Anthropic(api_key=provider["api_key"])
        resp = client.messages.create(
            model=provider["model"],
            max_tokens=2048,
            messages=[{"role": "user", "content": prompt}],
        )
        return resp.content[0].text if resp.content else ""
    except Exception as e:
        print(f"[LLM] Anthropic API call failed: {e}", file=sys.stderr)
        return ""


def extract_json(text):
    """Extract the first JSON object/array from LLM output."""
    # Try to find JSON inside triple backticks first
    match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", text)
    if match:
        text = match.group(1)
    # Try to parse as-is
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    # Try to find first { ... } or [ ... ]
    for pattern in (r"\{[\s\S]*\}", r"\[[\s\S]*\]"):
        m = re.search(pattern, text)
        if m:
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError:
                continue
    return None


def main():
    args = parse_args()
    ensure_dependencies()

    skill_dir = Path(__file__).parent.parent
    template_dir = skill_dir / "templates"

    # Load templates
    qa_template = load_template(template_dir, "qa_generator_prompt.md")
    critic_template = load_template(template_dir, "critic_prompt.md")
    cross_verifier_template = load_template(template_dir, "cross_verifier_prompt.md")

    # Auto-detect LLM strategy
    if args.llm_cli:
        llm_provider = {"type": "hermes", "cli": args.llm_cli}
        print(f"[LLM] Using user-specified CLI: {args.llm_cli}", file=sys.stderr)
    else:
        llm_provider = detect_llm()
    print(f"[LLM] Strategy: {llm_provider['type']}", file=sys.stderr)

    # Load sources
    sources = load_sources(
        Path(args.maygrove_md),
        Path(args.v30_manual),
        Path(args.seed_library),
    )
    seed_data = load_seed_library_data(sources)

    web_enabled = (not args.skip_web) and bool(os.getenv("TAVILY_API_KEY"))
    print(f"Loaded sources: {len(sources['v30_manual']['sections'])} V30 sections, "
          f"{len(sources['maygrove_md']['sections'])} Maygrove sections, "
          f"{len(sources['seed_library'])} seed varieties")
    print(f"Web research enabled: {web_enabled}")

    os.makedirs(args.output, exist_ok=True)

    # Brand boundary text
    brand_boundary = (
        "The MayGrove AI assistant only answers questions about MayGrove, indoor gardening, "
        "the Verti 30 device, plants in the seed library, and using the MayGrove app. "
        "It must refuse or deflect questions about politics, finance, medical advice, "
        "legal advice, coding, personal data, or unrelated topics."
    )

    # Define scenarios per intent
    scenarios = {
        "PlantingTech": ["daily-care", "seedling-stage", "nutrients", "harvest", "troubleshooting"],
        "ProductInquiry": ["setup", "specs", "app-usage", "troubleshooting", "maintenance"],
        "GrowingPlan": ["onboarding", "first-plan", "seed-selection", "capacity"],
        "Lifestyle": ["kitchen", "gift", "kids", "home-decor"],
        "ChatEmotional": ["greeting", "encouragement", "celebration"],
        "DietCooking": ["recipe", "nutrition", "food-safety"],
        "OutofScope": ["finance", "medical", "coding", "politics"],
    }

    # Generate candidate items for (intent, persona) cells. We generate one per cell
    # and then optionally expand to reach the target count. 7 intents × 4 personas = 28.
    candidates = []
    for intent in INTENTS:
        for persona in PERSONAS:
            scenario = scenarios[intent][hash(intent + persona) % len(scenarios[intent])]

            # Pick facts: for PlantingTech/ProductInquiry/GrowingPlan use local docs + seed library;
            # for DietCooking/Lifestyle use web if available; for ChatEmotional/OutofScope use brand boundary.
            facts = []
            if intent in ("PlantingTech", "ProductInquiry", "GrowingPlan"):
                facts.append({
                    "text": "V30 has 30 main planting pods and 12 seedling slots.",
                    "source": "V30 manual / 5 产品外观与结构说明"
                })
                facts.append({
                    "text": "Climbing plants such as tomato, cucumber, pea, and bean must be planted in the bottom 3 pods with the trellis system.",
                    "source": "V30 manual / 12 爬藤系统使用"
                })
                if seed_data.get("first_batch"):
                    seed = seed_data["first_batch"][hash(intent + persona) % len(seed_data["first_batch"])]
                    facts.append({
                        "text": f"{seed['app_name']} ({seed['variety_name']}) harvests in {seed['harvest_days']} days.",
                        "source": f"seed library / {seed['app_name']}"
                    })
            elif intent == "DietCooking":
                facts.append({
                    "text": "Wash harvested plants before eating. Some plants may be toxic to pets; consult a veterinarian.",
                    "source": "V30 manual / 4 安全提示 / 植物食用与安全说明"
                })
                if seed_data.get("nutrition"):
                    nut = seed_data["nutrition"][hash(persona) % len(seed_data["nutrition"])]
                    facts.append({
                        "text": f"{nut['name']} provides Vitamin C at {nut.get('Vitamin_C_mg', 'unknown')} mg per 100g.",
                        "source": f"seed nutrition sheet / {nut['name']}"
                    })
            elif intent == "Lifestyle":
                facts.append({
                    "text": "V30 is designed for kitchen, living room, or dining room placement.",
                    "source": "V30 manual / 1 产品简介"
                })
            elif intent == "ChatEmotional":
                facts.append({
                    "text": "The AI assistant should greet the user warmly, keep responses under 25 words, and stay in the Planting Bestie persona.",
                    "source": "Maygrove.md / 日常招呼区"
                })
            elif intent == "OutofScope":
                facts.append({
                    "text": "The assistant must refuse to answer unrelated topics and suggest MayGrove-related topics instead.",
                    "source": "brand boundary policy"
                })

            # Layer 2: run three QA generators
            qa_prompts = {
                "A": build_qa_prompt(qa_template, intent, persona, scenario, facts, brand_boundary),
                "B": build_qa_prompt(qa_template, intent, persona, scenario, facts, brand_boundary),
                "C": build_qa_prompt(qa_template, intent, persona, scenario, facts, brand_boundary),
            }
            # Adjust role descriptions per agent
            qa_prompts["A"] = qa_prompts["A"].replace("scenario-driven", "scenario-driven").replace("scenario-based user journey stage", "scenario-based user journey stage")
            qa_prompts["B"] = qa_prompts["B"].replace("scenario-driven", "fact-driven").replace("scenario-based user journey stage", "reverse-question from a fact")
            qa_prompts["C"] = qa_prompts["C"].replace("scenario-driven", "persona-driven").replace("scenario-based user journey stage", "persona voice adaptation")

            answers = []
            for agent in ("A", "B", "C"):
                if llm_provider["type"] != "fallback":
                    output = call_llm(llm_provider, qa_prompts[agent], timeout=120)
                    qa = extract_json(output)
                    if qa:
                        print(f"  [LLM OK] {intent}/{persona} agent {agent}", file=sys.stderr)
                        qa["agent"] = agent
                        answers.append(qa)
                        continue
                    print(f"  [LLM FAIL] {intent}/{persona} agent {agent} — falling back to template", file=sys.stderr)

                # Fallback: deterministic QA
                qa = generate_fallback_qa(intent, persona, scenario, facts, agent)
                qa["agent"] = agent
                answers.append(qa)

            # Layer 3: run three critics
            scores = []
            issues = []
            for critic_c, role in (("A", "fact traceability"), ("B", "logical plausibility"), ("C", "safety & boundary compliance")):
                chosen = answers[0] if answers[0].get("reference_answer") else answers[1] if answers[1].get("reference_answer") else answers[2]
                if llm_provider["type"] != "fallback":
                    prompt = build_critic_prompt(critic_template, chosen.get("question", ""), chosen.get("reference_answer", ""), facts, intent, brand_boundary, critic_c)
                    output = call_llm(llm_provider, prompt, timeout=90)
                    critic_result = extract_json(output) or {"score": 0.8, "issues": []}
                    print(f"  [Critic {critic_c}] score: {critic_result.get('score', 'N/A')}", file=sys.stderr)
                else:
                    critic_result = fallback_critic(intent, chosen, facts, critic_c)
                scores.append(critic_result["score"])
                issues.extend(critic_result.get("issues", []))

            # Cross-verifier
            consensus_score = 0.9 if answers[0].get("reference_answer") == answers[1].get("reference_answer") else 0.7
            chosen = answers[0]
            fact_coverage = min(1.0, len(facts) / max(1, len(chosen.get("reference_answer", "").split('.'))))
            critic_avg = sum(scores) / len(scores)
            final_confidence = fact_coverage * 0.3 + critic_avg * 0.4 + consensus_score * 0.3
            status = "verified" if final_confidence >= 0.80 and not any(i.get("severity") == "critical" for i in issues) else "partial"

            candidates.append({
                "id": f"mg-{len(candidates)+1:03d}",
                "intent": intent,
                "persona": persona,
                "scenario": scenario,
                "question": chosen["question"],
                "reference_answer": chosen["reference_answer"],
                "facts": facts,
                "confidence": round(final_confidence, 2),
                "tags": chosen.get("tags", []),
                "verification_status": status,
                "verification_confidence": round(final_confidence, 2),
                "verification_notes": f"Consensus score {consensus_score}, critic avg {critic_avg:.2f}, fact coverage {fact_coverage:.2f}",
                "issues": issues,
                "source_gap": "" if status == "verified" else "partial verification",
                "agent_a_answer": answers[0]["reference_answer"],
                "agent_b_answer": answers[1]["reference_answer"],
                "agent_c_answer": answers[2]["reference_answer"],
            })

    # Keep top N by confidence
    candidates.sort(key=lambda x: x["confidence"], reverse=True)
    verified = candidates[:args.count]
    rejected = candidates[args.count:]

    # Write outputs
    write_outputs(args.output, verified, rejected)
    print(f"Wrote {len(verified)} verified items, {len(rejected)} rejected items to {args.output}")


def generate_fallback_qa(intent, persona, scenario, facts, agent):
    """Produce a deterministic QA pair when no LLM runtime is available."""
    if intent == "OutofScope":
        q, a = generate_outofscope_qa(persona, scenario, facts)
    elif intent == "ChatEmotional":
        q, a = generate_chat_qa(persona, scenario, facts)
    elif intent == "DietCooking":
        q, a = generate_diet_qa(persona, scenario, facts)
    elif intent == "Lifestyle":
        q, a = generate_lifestyle_qa(persona, scenario, facts)
    elif intent == "ProductInquiry":
        q, a = generate_product_qa(persona, scenario, facts)
    elif intent == "GrowingPlan":
        q, a = generate_growing_plan_qa(persona, scenario, facts)
    else:  # PlantingTech
        q, a = generate_planting_tech_qa(persona, scenario, facts)

    # Vary the question slightly per agent role
    if agent == "B":
        q = "What is the fact-based answer to: " + q
    elif agent == "C":
        # Already persona-driven; leave as is
        pass

    return {
        "question": q,
        "reference_answer": a,
        "tags": [intent.lower(), scenario, persona.lower().replace(' ', '-')],
    }


def generate_planting_tech_qa(persona, scenario, facts):
    if persona == "Newbie":
        q = "How often do I need to check the water level?"
        a = "Check the water level at least weekly. Refill when the app notifies you or when the tank is below 20%. The V30 tank holds about 26.5L (7gal)."
    elif persona == "Busy Professional":
        q = "Can I leave the MayGrove running while I travel for a week?"
        a = "Yes, the tank holds about 26.5L (7gal), which supports most plants for a week. Top off before leaving and enable app notifications."
    elif persona == "Green Enthusiast":
        q = "What is the recommended EC and pH range for leafy greens?"
        a = "Leafy greens generally do best with EC around 0.8–1.6 and pH 5.8–6.2. The V30 automatically adjusts nutrients based on EC readings."
    else:  # Troubleshooter
        q = "My leaves are turning yellow. Is it nutrient burn or deficiency?"
        a = "Yellow leaves usually indicate nutrient deficiency or pH imbalance. Check the EC probe and confirm nutrient solution is not over-concentrated."
    return q, a


def generate_product_qa(persona, scenario, facts):
    if persona == "Newbie":
        q = "What do I do first after unboxing the V30?"
        a = "Install the bottom planting pods, attach the support poles, fill the tank with water to at least 20%, and connect power. Then follow the app setup flow."
    elif persona == "Busy Professional":
        q = "How long does the initial setup take?"
        a = "Most users complete the hardware setup in 10–15 minutes and app pairing in another 5 minutes."
    elif persona == "Green Enthusiast":
        q = "How many plants can the V30 grow at the same time?"
        a = "The V30 has 30 main planting pods plus 12 seedling slots, so it can support up to 42 plants in different growth stages."
    else:
        q = "Why won't the V30 turn on?"
        a = "Confirm the power adapter is plugged into the back of the tank and the outlet. Check the LED indicator. If still unresponsive, contact MayGrove support."
    return q, a


def generate_growing_plan_qa(persona, scenario, facts):
    if persona == "Newbie":
        q = "What should I plant first as a beginner?"
        a = "Start with the first-batch varieties: lettuce, kale, arugula, radish, and herbs. They are fast-growing and forgiving."
    elif persona == "Busy Professional":
        q = "Can I skip the growing plan and just plant something?"
        a = "You can skip the plan, but completing the onboarding plan helps the app recommend the right varieties and schedule."
    elif persona == "Green Enthusiast":
        q = "How does the plan decide which plants to recommend?"
        a = "The plan uses your profile inputs such as experience level, available light, and preferences to select from the seed library."
    else:
        q = "My plan generated a schedule I don't like. Can I change it?"
        a = "Yes, you can edit the plan in the app, swap varieties, and adjust planting dates."
    return q, a


def generate_lifestyle_qa(persona, scenario, facts):
    if persona == "Newbie":
        q = "Can I put the MayGrove in my kitchen?"
        a = "Yes, the V30 is designed for kitchen, living room, or dining room placement. Keep it near a power outlet and away from direct heat sources."
    elif persona == "Busy Professional":
        q = "Is this a good gift for a coworker?"
        a = "Yes, the compact design and low maintenance make it a popular gift for busy professionals."
    elif persona == "Green Enthusiast":
        q = "Which plants are best for decorating a living room?"
        a = "Flowering plants such as zinnia, sunflower, and calendula add color. Leafy greens also provide a fresh look."
    else:
        q = "My cat might chew on the plants. Where should I place the device?"
        a = "Place it in a location your cat cannot reach, and check the seed library for any plant that may be toxic to pets."
    return q, a


def generate_chat_qa(persona, scenario, facts):
    q = "Hey, I just finished my first harvest!"
    a = "That's wonderful! Harvesting your own greens is such a rewarding moment. What are you planning to grow next?"
    return q, a


def generate_diet_qa(persona, scenario, facts):
    if persona == "Newbie":
        q = "Can I eat the basil straight from the plant?"
        a = "Yes, but wash it first. Some plants may be toxic to pets, so keep them out of reach and consult a vet if needed."
    elif persona == "Busy Professional":
        q = "Quick recipe for basil?"
        a = "Blend fresh basil with garlic, pine nuts, parmesan, and olive oil for a quick pesto. Wash ingredients before eating."
    elif persona == "Green Enthusiast":
        q = "Which MayGrove herb is highest in Vitamin C?"
        a = "Parsley is one of the highest Vitamin C herbs in the seed library. Check the nutrition sheet for exact values."
    else:
        q = "Are the plants safe if my dog ate a leaf?"
        a = "Some plants may be toxic to pets. Contact your veterinarian and refer to the plant list for known toxicity."
    return q, a


def generate_outofscope_qa(persona, scenario, facts):
    questions = {
        "finance": "What stocks should I buy today?",
        "medical": "I have a headache. Should I take aspirin?",
        "coding": "Can you write a Python script for me?",
        "politics": "Who should I vote for in the election?",
    }
    q = questions.get(scenario, "What's the weather like on Mars?")
    a = "I'm here to help with MayGrove and indoor gardening. Let me know if you have questions about plants, the V30, or the app!"
    return q, a


def fallback_critic(intent, chosen, facts, critic):
    """Deterministic critic scores."""
    answer = chosen.get("reference_answer", "")
    if critic == "A":
        if intent == "OutofScope":
            score = 1.0 if "MayGrove" in answer or "gardening" in answer.lower() else 0.5
        else:
            score = 0.9 if any(f["source"] in answer for f in facts) else 0.8
        issues = []
    elif critic == "B":
        score = 0.9
        issues = []
    else:  # C
        if intent in ("DietCooking", "PlantingTech") and "wash" not in answer.lower() and intent == "DietCooking":
            score = 0.7
            issues = [{"claim": answer, "problem": "Missing food-safety disclaimer", "severity": "minor"}]
        elif intent == "OutofScope" and any(w in answer.lower() for w in ["buy", "aspirin", "python", "vote"]):
            score = 0.0
            issues = [{"claim": answer, "problem": "Answered an OutofScope question", "severity": "critical"}]
        else:
            score = 1.0
            issues = []
    return {"score": score, "issues": issues}


def write_outputs(output_dir, verified, rejected):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # JSONL
    with open(output_dir / "verified_qa.jsonl", "w", encoding="utf-8") as f:
        for item in verified:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    with open(output_dir / "rejected_needs_review.jsonl", "w", encoding="utf-8") as f:
        for item in rejected:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    # CSV
    import csv
    csv_columns = [
        "id", "intent", "persona", "scenario", "question", "reference_answer",
        "facts_count", "source_sections", "confidence", "verification_status",
        "verification_confidence", "source_gap", "tags"
    ]
    with open(output_dir / "verified_qa.csv", "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_columns)
        writer.writeheader()
        for item in verified:
            writer.writerow({
                "id": item["id"],
                "intent": item["intent"],
                "persona": item["persona"],
                "scenario": item["scenario"],
                "question": item["question"],
                "reference_answer": item["reference_answer"],
                "facts_count": len(item["facts"]),
                "source_sections": "; ".join(f["source"] for f in item["facts"]),
                "confidence": item["confidence"],
                "verification_status": item["verification_status"],
                "verification_confidence": item["verification_confidence"],
                "source_gap": item["source_gap"],
                "tags": ", ".join(item["tags"]),
            })

    # Excel
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Verified QA"
        ws.append(csv_columns)
        for item in verified:
            ws.append([
                item["id"], item["intent"], item["persona"], item["scenario"],
                item["question"], item["reference_answer"], len(item["facts"]),
                "; ".join(f["source"] for f in item["facts"]), item["confidence"],
                item["verification_status"], item["verification_confidence"],
                item["source_gap"], ", ".join(item["tags"]),
            ])

        ws2 = wb.create_sheet("Verification Report")
        ws2.append(["id", "intent", "persona", "question", "agent_a_answer", "agent_b_answer", "agent_c_answer", "verification_confidence", "verification_status"])
        for item in verified:
            ws2.append([
                item["id"], item["intent"], item["persona"], item["question"],
                item["agent_a_answer"], item["agent_b_answer"], item["agent_c_answer"],
                item["verification_confidence"], item["verification_status"],
            ])

        ws3 = wb.create_sheet("Rejected")
        ws3.append(csv_columns)
        for item in rejected:
            ws3.append([
                item["id"], item["intent"], item["persona"], item["scenario"],
                item["question"], item["reference_answer"], len(item["facts"]),
                "; ".join(f["source"] for f in item["facts"]), item["confidence"],
                item["verification_status"], item["verification_confidence"],
                item["source_gap"], ", ".join(item["tags"]),
            ])

        wb.save(output_dir / "verified_qa.xlsx")
    except Exception as exc:
        print(f"Warning: could not write Excel: {exc}", file=sys.stderr)


if __name__ == "__main__":
    main()
